"""
Excel Handler - Excel dosyalarini okuma ve yazma islemleri
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime


class ExcelHandler:
    """Excel dosyalarini yonetir."""
    
    def __init__(self, file_path: str = None):
        self.file_path = file_path
        self.workbook = None
        self.sheet = None
    
    def create_sample_data(self, output_path: str) -> str:
        """Demo icin ornek urun verileri olusturur."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Urunler"
        
        # Basliklar
        headers = ["Urun Adi", "Birim Fiyat (TL)", "Adet", "Toplam", "KDV (%18)", "Genel Toplam"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        # Ornek veriler
        products = [
            ("Laptop", 25000, 2),
            ("Mouse", 350, 10),
            ("Klavye", 750, 5),
            ("Monitor", 8500, 3),
            ("Kulaklik", 1200, 8),
            ("Webcam", 2100, 4),
            ("SSD 1TB", 3500, 6),
            ("RAM 16GB", 2800, 5),
        ]
        
        for row, (name, price, qty) in enumerate(products, 2):
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=price)
            ws.cell(row=row, column=3, value=qty)
            # Toplam, KDV ve Genel Toplam sutunlari bot tarafindan doldurulacak
            ws.cell(row=row, column=4, value="")  # Toplam
            ws.cell(row=row, column=5, value="")  # KDV
            ws.cell(row=row, column=6, value="")  # Genel Toplam
        
        # Sutun genislikleri
        widths = [20, 18, 10, 15, 15, 18]
        for i, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = width
        
        wb.save(output_path)
        print(f"[EXCEL] Ornek veri olusturuldu: {output_path}")
        return output_path
    
    def load(self, file_path: str = None) -> list:
        """Excel dosyasini yukler ve verileri dondurur."""
        path = file_path or self.file_path
        self.workbook = load_workbook(path)
        self.sheet = self.workbook.active
        self.file_path = path
        
        data = []
        headers = [cell.value for cell in self.sheet[1]]
        
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Bos satirlari atla
                item = dict(zip(headers, row))
                data.append(item)
        
        print(f"[EXCEL] {len(data)} urun yuklendi: {path}")
        return data
    
    def update_row(self, row_num: int, total: float, kdv: float, grand_total: float):
        """Belirli bir satiri gunceller."""
        if not self.sheet:
            raise ValueError("Once Excel dosyasi yuklenmeli!")
        
        self.sheet.cell(row=row_num, column=4, value=total)
        self.sheet.cell(row=row_num, column=5, value=kdv)
        self.sheet.cell(row=row_num, column=6, value=grand_total)
        
        # Sayi formati
        for col in [4, 5, 6]:
            self.sheet.cell(row=row_num, column=col).number_format = '#,##0.00 TL'
    
    def add_summary(self, grand_total: float):
        """Dosyanin sonuna ozet satiri ekler."""
        last_row = self.sheet.max_row + 2
        
        summary_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
        summary_font = Font(bold=True, color="FFFFFF", size=12)
        
        self.sheet.cell(row=last_row, column=1, value="GENEL TOPLAM")
        self.sheet.cell(row=last_row, column=1).font = summary_font
        self.sheet.cell(row=last_row, column=1).fill = summary_fill
        
        self.sheet.cell(row=last_row, column=6, value=grand_total)
        self.sheet.cell(row=last_row, column=6).font = summary_font
        self.sheet.cell(row=last_row, column=6).fill = summary_fill
        self.sheet.cell(row=last_row, column=6).number_format = '#,##0.00 TL'
        
        # Timestamp
        self.sheet.cell(row=last_row + 1, column=1, 
                       value=f"Islem Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    
    def save(self, output_path: str = None):
        """Degisiklikleri kaydeder."""
        path = output_path or self.file_path
        self.workbook.save(path)
        print(f"[EXCEL] Kaydedildi: {path}")


if __name__ == "__main__":
    # Test
    handler = ExcelHandler()
    handler.create_sample_data("test_input.xlsx")
