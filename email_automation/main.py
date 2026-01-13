"""
PDF Rapor Botu - Ana Orkestrator
================================
Excel'den veri okuyup toplu PDF rapor uretir.

Kullanim:
    python main.py
    python main.py demo
"""

import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, str(Path(__file__).parent))
from pdf_generator import PDFReportGenerator


INPUT_FILE = "data/urunler.xlsx"
OUTPUT_DIR = "output/reports"


def create_sample_excel(filepath: str):
    """Ornek urun verileri olusturur."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Urunler"
    
    headers = ["Urun Adi", "Adet", "Birim Fiyat (TL)"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    products = [
        ("Laptop", 2, 25000),
        ("Mouse", 10, 350),
        ("Klavye", 5, 750),
        ("Monitor", 3, 8500),
        ("Kulaklik", 8, 1200),
    ]
    
    for row, (name, qty, price) in enumerate(products, 2):
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=qty)
        ws.cell(row=row, column=3, value=price)
    
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"[EXCEL] Ornek veri olusturuldu: {filepath}")


def load_products(filepath: str) -> list:
    """Excel'den urunleri yukler."""
    wb = load_workbook(filepath)
    ws = wb.active
    
    data = []
    headers = [cell.value for cell in ws[1]]
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            data.append(dict(zip(headers, row)))
    
    print(f"[EXCEL] {len(data)} urun yuklendi")
    return data


def print_banner():
    print("\n" + "=" * 50)
    print(">>> PDF RAPOR OLUSTURUCU <<<")
    print("=" * 50)
    print(f"    Baslangic: {datetime.now().strftime('%H:%M:%S')}")
    print("=" * 50 + "\n")


def main():
    print_banner()
    
    input_path = Path(__file__).parent / INPUT_FILE
    
    # Dosya yoksa ornek olustur
    if not input_path.exists():
        create_sample_excel(str(input_path))
    
    # Verileri yukle
    print("[ADIM 1/2] Excel'den veri yukleniyor...")
    print("-" * 40)
    products = load_products(str(input_path))
    
    # PDF olustur
    print("\n[ADIM 2/2] PDF rapor olusturuluyor...")
    print("-" * 40)
    
    output_path = Path(__file__).parent / OUTPUT_DIR
    generator = PDFReportGenerator(str(output_path))
    
    pdf_path = generator.create_sales_report(
        title="Urun Satis Raporu",
        data=products,
        company_name="RPA Demo Sirket"
    )
    
    # Ozet
    print("\n" + "=" * 50)
    print(">>> ISLEM TAMAMLANDI! <<<")
    print("=" * 50)
    print(f"    PDF: {pdf_path}")
    print("=" * 50 + "\n")


def demo():
    """Hizli demo."""
    print("\n=== DEMO MOD ===\n")
    
    output_path = Path(__file__).parent / OUTPUT_DIR
    generator = PDFReportGenerator(str(output_path))
    
    demo_data = [
        {"urun": "iPhone 15", "adet": 5, "fiyat": 65000},
        {"urun": "MacBook Pro", "adet": 2, "fiyat": 95000},
        {"urun": "AirPods Pro", "adet": 10, "fiyat": 8500},
        {"urun": "iPad Air", "adet": 3, "fiyat": 32000},
    ]
    
    generator.create_sales_report(
        title="Demo Satis Raporu",
        data=demo_data,
        filename="demo_rapor.pdf",
        company_name="Apple Store TR"
    )
    
    print("\n[OK] Demo tamamlandi!\n")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "demo":
        demo()
    else:
        main()
