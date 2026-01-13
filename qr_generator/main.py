"""
QR Kod Olusturucu - Ana Bot
===========================
Excel'den veri okuyup toplu QR kod uretir.

Kullanim:
    python main.py
"""

import sys
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Modul yolunu ayarla
sys.path.insert(0, str(Path(__file__).parent))

from qr_local import create_qr_code, create_social_media_qr, create_vcard_qr, create_wifi_qr
from qr_api import create_qr_via_api, create_social_qr_api


# ============================================================
# AYARLAR
# ============================================================
INPUT_FILE = "data/social_links.xlsx"
OUTPUT_DIR = "output/qr_codes"
USE_API = False  # True = API kullan, False = Yerel kutuphane


def create_sample_excel(filepath: str):
    """Ornek sosyal medya verileri olusturur."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sosyal Medya"
    
    # Basliklar
    headers = ["Platform", "Kullanici Adi", "Aciklama"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Ornek veriler - Kullanici kendi bilgilerini girecek
    sample_data = [
        ("Instagram", "kullanici_adi", "Instagram profilim"),
        ("LinkedIn", "kullanici-adi", "LinkedIn profilim"),
        ("GitHub", "kullanici", "GitHub profilim"),
        ("Twitter", "kullanici", "Twitter/X profilim"),
        ("YouTube", "kanal_adi", "YouTube kanalim"),
        ("WhatsApp", "905551234567", "WhatsApp numaram"),
    ]
    
    for row, (platform, username, desc) in enumerate(sample_data, 2):
        ws.cell(row=row, column=1, value=platform)
        ws.cell(row=row, column=2, value=username)
        ws.cell(row=row, column=3, value=desc)
    
    # Sutun genislikleri
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    
    Path(filepath).parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"[EXCEL] Ornek veri olusturuldu: {filepath}")
    print("[INFO] Lutfen Excel'i acip kendi bilgilerinizi girin!")


def load_social_data(filepath: str) -> list:
    """Excel'den sosyal medya verilerini yukler."""
    wb = load_workbook(filepath)
    ws = wb.active
    
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:  # Platform ve kullanici adi varsa
            data.append({
                "platform": str(row[0]).strip(),
                "username": str(row[1]).strip(),
                "description": str(row[2]).strip() if row[2] else ""
            })
    
    print(f"[EXCEL] {len(data)} kayit yuklendi")
    return data


def print_banner():
    """Baslangic banner'i."""
    print("\n" + "=" * 50)
    print(">>> QR KOD OLUSTURUCU BOT <<<")
    print("    Sosyal Medya & Toplu QR Uretimi")
    print("=" * 50)
    print(f"    Baslangic: {datetime.now().strftime('%H:%M:%S')}")
    print(f"    Mod: {'API' if USE_API else 'Yerel Kutuphane'}")
    print("=" * 50 + "\n")


def main():
    print_banner()
    
    # Cikti klasorunu olustur
    Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)
    
    # 1. EXCEL - Veri Yukleme
    print("[ADIM 1/3] Excel'den veri yukleniyor...")
    print("-" * 40)
    
    input_path = Path(__file__).parent / INPUT_FILE
    
    # Dosya yoksa ornek olustur
    if not input_path.exists():
        create_sample_excel(str(input_path))
        print("\n[UYARI] Ornek Excel olusturuldu.")
        print("[UYARI] Lutfen data/social_links.xlsx dosyasini")
        print("[UYARI] kendi bilgilerinizle doldurup tekrar calistirin.\n")
        return
    
    social_data = load_social_data(str(input_path))
    
    if not social_data:
        print("[HATA] Excel'de veri bulunamadi!")
        return
    
    # 2. QR URETIMI
    print("\n[ADIM 2/3] QR kodlar uretiliyor...")
    print("-" * 40)
    
    generated = []
    output_path = Path(__file__).parent / OUTPUT_DIR
    
    for i, item in enumerate(social_data, 1):
        platform = item["platform"]
        username = item["username"]
        
        print(f"\n[{i}/{len(social_data)}] {platform}: {username}")
        
        try:
            if USE_API:
                filepath = create_social_qr_api(
                    platform=platform,
                    username=username,
                    output_dir=str(output_path)
                )
            else:
                filepath = create_social_media_qr(
                    platform=platform,
                    username=username,
                    output_dir=str(output_path)
                )
            
            if filepath:
                generated.append({
                    "platform": platform,
                    "username": username,
                    "file": filepath
                })
        except Exception as e:
            print(f"[HATA] {platform} QR olusturulamadi: {e}")
    
    # 3. OZET
    print("\n" + "=" * 50)
    print(">>> ISLEM TAMAMLANDI! <<<")
    print("=" * 50)
    print(f"    Olusturulan QR: {len(generated)}/{len(social_data)}")
    print(f"    Cikti Klasoru: {output_path}")
    print("=" * 50)
    
    # Olusturulan dosyalar
    print("\nOlusturulan QR Kodlar:")
    for item in generated:
        print(f"  - {item['platform']}: {item['file']}")
    
    print()


def demo():
    """Hizli demo - ornek QR'lar olusturur."""
    print("\n=== DEMO MOD ===\n")
    
    output_path = Path(__file__).parent / OUTPUT_DIR
    Path(output_path).mkdir(parents=True, exist_ok=True)
    
    # 1. Basit URL QR
    print("[1] URL QR Kodu...")
    create_qr_code(
        data="https://github.com",
        filename="demo_github_url",
        output_dir=str(output_path)
    )
    
    # 2. Sosyal medya QR'lari
    print("\n[2] Sosyal Medya QR Kodlari...")
    platforms = ["instagram", "linkedin", "github", "youtube"]
    for platform in platforms:
        create_social_media_qr(
            platform=platform,
            username="demo_user",
            output_dir=str(output_path)
        )
    
    # 3. vCard QR
    print("\n[3] Kartvizit (vCard) QR Kodu...")
    create_vcard_qr(
        name="Demo Kullanici",
        phone="+905551234567",
        email="demo@example.com",
        company="Demo Sirket",
        title="RPA Developer",
        output_dir=str(output_path)
    )
    
    # 4. WiFi QR
    print("\n[4] WiFi QR Kodu...")
    create_wifi_qr(
        ssid="Demo_WiFi",
        password="demo1234",
        output_dir=str(output_path)
    )
    
    print("\n" + "=" * 40)
    print("[OK] Demo tamamlandi!")
    print(f"QR kodlar: {output_path}")
    print("=" * 40 + "\n")


if __name__ == "__main__":
    import sys
    
    if len(sys.argv) > 1 and sys.argv[1] == "demo":
        demo()
    else:
        main()
