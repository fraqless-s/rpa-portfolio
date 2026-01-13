"""
RPA Project 1: Multi-Site E-Ticaret Fiyat Takip Botu
====================================================
Bu bot, Amazon, Trendyol, Hepsiburada ve demo sitelerinden
ürün bilgilerini çeker ve Excel'e kaydeder.

Desteklenen Siteler:
- Amazon Türkiye (amazon.com.tr)
- Trendyol (trendyol.com)
- Hepsiburada (hepsiburada.com)
- Books to Scrape (books.toscrape.com - Demo)
"""

import asyncio
import re
from abc import ABC, abstractmethod
from datetime import datetime
from urllib.parse import urlparse
from playwright.async_api import async_playwright, Page
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================
# AYARLAR - Takip edilecek ürünler
# ============================================================
PRODUCTS_TO_TRACK = [
    {
        "name": "iPhone 15 256GB (Amazon)",
        "url": "https://www.amazon.com.tr/Apple-iPhone-15-256-GB/dp/B0CHX43FKD/",
    },
    {
        "name": "iPhone 15 256GB (Hepsiburada)",
        "url": "https://www.hepsiburada.com/apple-iphone-15-256-gb-siyah-p-HBCV00004X9ZCN",
    },
    {
        "name": "iPhone 15 256GB (Trendyol)",
        "url": "https://www.trendyol.com/apple/iphone-15-256-gb-siyah-p-762254844",
    },
]

# ============================================================
# SCRAPER CLASSES
# ============================================================

class BaseScraper(ABC):
    """Tüm scraper'lar için temel sınıf."""
    
    @abstractmethod
    async def scrape(self, page: Page) -> dict:
        """Sayfadan veriyi çeker ve sözlük döndürür."""
        pass

    async def get_text(self, page: Page, selector: str) -> str:
        """Element metnini güvenli bir şekilde alır."""
        try:
            element = await page.wait_for_selector(selector, timeout=3000)
            if element:
                return await element.inner_text()
        except:
            pass
        return None

    def clean_price(self, price_text: str) -> float:
        """Fiyat metnini sayıya çevirir."""
        if not price_text:
            return 0.0
        # Sadece sayıları ve virgül/noktayı bırak
        clean = re.sub(r'[^\d,.]', '', price_text)
        # Türkçe formatı (1.299,90) -> (1299.90) standardına çevir
        if ',' in clean and '.' in clean:
            clean = clean.replace('.', '').replace(',', '.')
        elif ',' in clean:
            clean = clean.replace(',', '.')
        
        try:
            return float(clean)
        except ValueError:
            return 0.0

class AmazonScraper(BaseScraper):
    async def scrape(self, page: Page) -> dict:
        title = await self.get_text(page, "#productTitle")
        
        # Amazon fiyatları bazen whole/fraction olarak ayrılır
        price_whole = await self.get_text(page, ".a-price-whole")
        price_fraction = await self.get_text(page, ".a-price-fraction")
        
        if price_whole:
            price_text = f"{price_whole}"
            if price_fraction:
                price_text +=f",{price_fraction}"
        else:
            # Alternatif fiyat alanı
            price_text = await self.get_text(page, "#priceblock_ourprice") or \
                         await self.get_text(page, "#apex_desktop .a-offscreen")
            
        availability = await self.get_text(page, "#availability")
        
        return {
            "title": title.strip() if title else "Başlık Bulunamadı",
            "price": price_text.strip() if price_text else "N/A",
            "stock": availability.strip().replace("\n", "") if availability else "Belirsiz"
        }

class TrendyolScraper(BaseScraper):
    async def scrape(self, page: Page) -> dict:
        title = await self.get_text(page, "h1.product-name") or await self.get_text(page, ".product-name-text")
        price = await self.get_text(page, ".product-price-container") or await self.get_text(page, ".prc-dsc")
        
        # Trendyol stok kontrolü biraz dolaylıdır, sepete ekle butonu var mı?
        add_to_cart = await page.query_selector(".add-to-basket-button-text")
        stock = "Stokta Var" if add_to_cart else "Stok Yok/Tükendi"

        return {
            "title": title.strip() if title else "Başlık Bulunamadı",
            "price": price.strip() if price else "N/A",
            "stock": stock
        }

class HepsiburadaScraper(BaseScraper):
    async def scrape(self, page: Page) -> dict:
        title = await self.get_text(page, "h1#product-name")
        price = await self.get_text(page, '[data-test-id="price-current-price"]')
        
        # Hepsiburada stok/kargo bilgisi
        stock_info = await self.get_text(page, ".shipping-date") or \
                     await self.get_text(page, ".delivery-info") or \
                     await self.get_text(page, ".product-inventory-status")

        return {
            "title": title.strip() if title else "Başlık Bulunamadı",
            "price": price.strip() if price else "N/A",
            "stock": stock_info.strip() if stock_info else "Stok Bilgisi Yok"
        }

class LegacyScraper(BaseScraper):
    """books.toscrape.com için"""
    async def scrape(self, page: Page) -> dict:
        title = await self.get_text(page, "h1")
        price = await self.get_text(page, "p.price_color")
        stock = await self.get_text(page, "p.availability")
        
        return {
            "title": title.strip() if title else "Not Found",
            "price": price.strip() if price else "N/A",
            "stock": stock.strip() if stock else "N/A"
        }

def get_scraper(url: str) -> BaseScraper:
    domain = urlparse(url).netloc
    if "amazon" in domain:
        return AmazonScraper()
    elif "trendyol" in domain:
        return TrendyolScraper()
    elif "hepsiburada" in domain:
        return HepsiburadaScraper()
    else:
        return LegacyScraper()

# ============================================================
# MAIN LOGIC
# ============================================================

async def process_product(page: Page, product_info: dict) -> dict:
    url = product_info["url"]
    print(f"[SCRAPING] {product_info['name']} -> {url}")
    
    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=45000)
        
        # Bot algılamayı önlemek için rastgele bekleme
        await page.wait_for_timeout(2000)
        
        scraper = get_scraper(url)
        data = await scraper.scrape(page)
        
        # Fiyat sayısal çevrimi
        price_numeric = scraper.clean_price(data["price"])
        
        return {
            "name": product_info["name"],
            "title": data["title"],
            "price": data["price"],
            "price_numeric": price_numeric,
            "availability": data["stock"],
            "url": url,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "status": "[OK] Başarılı"
        }
        
    except Exception as e:
        print(f"[ERROR] {product_info['name']} Hatası: {str(e)}")
        return {
            "name": product_info["name"],
            "title": "HATA",
            "price": "N/A",
            "price_numeric": 0,
            "availability": "N/A",
            "url": url,
            "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "status": f"[ERROR] {str(e)[:50]}"
        }

def save_to_excel(products_data: list, filename: str = "fiyat_takibi.xlsx"):
    print(f"\n[EXCEL] Kaydediliyor: {filename}")
    wb = Workbook()
    ws = wb.active
    ws.title = "Fiyatlar"
    
    headers = ["Ürün Adı", "Site Başlığı", "Fiyat (Metin)", "Fiyat (Sayı)", "Stok/Durum", "Link", "Zaman", "Durum"]
    
    # Stil
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        
    for row, p in enumerate(products_data, 2):
        ws.cell(row=row, column=1, value=p["name"])
        ws.cell(row=row, column=2, value=p["title"])
        ws.cell(row=row, column=3, value=p["price"])
        ws.cell(row=row, column=4, value=p["price_numeric"])
        ws.cell(row=row, column=5, value=p["availability"])
        ws.cell(row=row, column=6, value=p["url"])
        ws.cell(row=row, column=7, value=p["timestamp"])
        ws.cell(row=row, column=8, value=p["status"])

    # Genişlik ayarla
    dims = [20, 50, 15, 12, 25, 60, 20, 30]
    for i, w in enumerate(dims, 1):
        ws.column_dimensions[chr(64 + i)].width = w
        
    wb.save(filename)
    print(f"[OK] Kayıt tamamlandı.")

async def main():
    print("Bot Başlatılıyor...")
    async with async_playwright() as p:
        # Görünür modda başlat (headless=False)
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
        )
        page = await context.new_page()
        
        results = []
        for product in PRODUCTS_TO_TRACK:
            res = await process_product(page, product)
            results.append(res)
            
        await browser.close()
        
    save_to_excel(results)

if __name__ == "__main__":
    asyncio.run(main())
